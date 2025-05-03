# -*- coding: utf-8 -*-
"""Para facilitar o uso do shareplum."""
from __future__ import annotations

import io
import pickle
import requests
import openpyxl

import pandas as pd

from infra_copel.sharepoint.shareplum import Office365
from infra_copel.sharepoint.shareplum import Site
from infra_copel.sharepoint.shareplum.site import Version
from infra_copel.sharepoint.shareplum.folder import _Folder
from infra_copel.sharepoint.shareplum.errors import ShareplumRequestError

from infra_copel.sharepoint.errors import _analyze_shareplum_err
from infra_copel.sharepoint.errors import LockedError, NotFoundError


class SharepointSite:
    """Classe que representa um site do Sharepoint."""

    def __init__(
        self,
        share_point_site: str,
        username: str,
        password: str,
        site_url: str,
        verify_ssl: bool,
    ):
        """
        Construtor para acesso a arquivos compartilhados via Sharepoint.

        Parameters
        ----------
        share_point_site : str
            Site do sharepoint ("https://<something>.sharepoint.com/)
        username : str
            Usuário do sharepoint.
        password : str
            Senha do usuário do sharepoint.
        site_url : str
            URL completa do site
            ("https://<something>.sharepoint.com/sites/something).
        verify_ssl : bool, default is True
            Verificar SSL. Desabilitar se houver problemas com certificados.

        """
        if verify_ssl is False:
            # Desabilita warnings
            # pylint:disable=no-member
            requests.packages.urllib3.disable_warnings(
                requests.packages.urllib3.exceptions.InsecureRequestWarning
            )

        # Conexão com o sharepoint
        try:
            authcookie = Office365(
                share_point_site=share_point_site, username=username, password=password
            ).get_cookies()
        except ShareplumRequestError as err:
            _analyze_shareplum_err(err)

        # Site desejado
        self._site = Site(
            site_url=site_url,
            version=Version.v2016,
            authcookie=authcookie,
            verify_ssl=verify_ssl,
        )

    def get_folder(self, folder: str) -> _Folder:
        """
        Retorna uma pasta do site do sharepoint.

        Cria recursivamente a pasta caso não exista.

        Parameters
        ----------
        folder : str
            Endereço da pasta.

        Returns
        -------
        _Folder
            Objeto que representa uma pasta do Sharepoint.
        """
        try:
            _folder = self._site.Folder(folder)
        except ShareplumRequestError:
            # Provavelmente as pastas não existem
            # Cria de forma recursiva
            subpastas = folder.split("/")
            for i in range(1, len(subpastas) + 1):
                _folder = self._site.Folder("/".join(subpastas[:i]))
        return _folder

    def get_file(self, folder: str, filename: str):
        """
        Retorna conteúdo de um arquivo do sharepoint.

        Parameters
        ----------
        folder : str
            Endereço da pasta.
        filename : str
            Nome do arquivo.

        Returns
        -------
        bytes
            Conteúdo do arquivo lido.

        """
        _folder = self.get_folder(folder)

        try:
            return _folder.get_file(filename)
        except ShareplumRequestError as err:
            _analyze_shareplum_err(err)

    def check_in(
        self,
        folder: str,
        filename: str,
        comment: str,
    ) -> None:
        """
        Realiza o check in do arquivo.

        Parameters
        ----------
        folder : str
            Endereço da pasta.
        filename : str
            Nome do arquivo.
        comment : str
            Comentário a ser inserido no check_in.

        """
        _folder = self.get_folder(folder)

        try:
            _folder.check_in(filename, comment)
        except ShareplumRequestError as err:
            _analyze_shareplum_err(err)

    def check_out(
        self,
        folder: str,
        filename: str,
    ) -> None:
        """
        Realiza o check out do arquivo.

        Parameters
        ----------
        folder : str
            Endereço da pasta.
        filename : str
            Nome do arquivo.

        """
        _folder = self.get_folder(folder)
        try:
            _folder.check_out(filename)
        except ShareplumRequestError as err:
            _analyze_shareplum_err(err)

    def check_out_forced(
        self,
        folder: str,
        filename: str,
        comment: str = "forced by script",
    ) -> None:
        """
        Realiza o check out de maneira forçada (tenta check-in se necessário).

        Parameters
        ----------
        folder : str
            Endereço da pasta.
        filename : str
            Nome do arquivo.
        comment : str
            Comentário a ser inserido no check_in.

        """
        try:
            self.check_out(folder, filename)
        except LockedError:
            self.check_in(folder, filename, comment)
            self.check_out(folder, filename)
        except NotFoundError:
            pass

    def list_filenames(
        self,
        folder: str,
    ) -> list[str]:
        """
        Lista nomes dos arquivos da pasta

                Parameters
                ----------
                folder : str
                    Endereço da pasta.
        """
        _folder = self.get_folder(folder)

        files = [f["Name"] for f in _folder.files]

        return sorted(files)

    def upload_file(
        self,
        folder: str,
        content: str | bytes | io.BytesIO,
        filename: str,
        check_in_msg: str = "upload_file",
        delete_old_first: bool = True,
    ) -> None:
        """
        Realiza o upload do conteúdo do arquivo.

        Parameters
        ----------
        folder : str
            Endereço da pasta.
        content : str | bytes | io.BytesIO
            Conteúdo do arquivo.
        filename : str
            Nome do arquivo.

        """
        _folder = self.get_folder(folder)

        if isinstance(content, io.BytesIO):
            # Retorna o BytesIO para o início e realiza a leitura
            content.seek(0)
            content = content.read()

        if delete_old_first:
            try:
                _folder.delete_file(filename)
            except Exception:
                # Não importa se deu algum erro
                pass
        
        # Envia o arquivo
        _folder.upload_file(content, filename)

        try:
            self.check_in(folder, filename, comment=check_in_msg)
        except LockedError:
            # Não estava em check out
            pass

    def upload_file_forced(
        self,
        folder: str,
        content: str | bytes | io.BytesIO,
        filename: str,
        check_out: bool = True,
    ) -> None:
        """
        Realiza o upload do conteúdo do arquivo mesmo com outro checkout.

        Parameters
        ----------
        folder : str
            Endereço da pasta.
        content : str | bytes | io.BytesIO
            Conteúdo do arquivo.
        filename : str
            Nome do arquivo.
        check_out: bool, default True
            Realizar check out após o upload

        """
        self.check_out_forced(folder, filename, comment="upload_file_forced")
        self.upload_file(folder, content, filename)
        if check_out:
            self.check_out(folder, filename)

    def write_to_pickle(
        self,
        obj: object,
        folder: str,
        filename: str,
    ) -> None:
        """
        Salva o objeto como pickle no sharepoint.

        Parameters
        ----------
        save_object : object
            Objeto a ser salvo.
        folder : str
            Pasta onde será salvo.
        filename : str
            Nome do arquivo pickle de destino.

        Returns
        -------
        None

        """
        # Cria um objeto binário
        bytes_io = io.BytesIO()
        # Salva o conteúdo neste objeto
        pickle.dump(obj, bytes_io)
        # Realiza o upload do conteúdo
        return self.upload_file(folder, bytes_io, filename)

    def read_pickle(
        self,
        folder: str,
        filename: str,
    ) -> object:
        """
        Ler um pickle salvo no sharepoint e retorna um objeto

        Parameters
        ----------
        folder : str
            Pasta onde o pickle está localizado
        filename : str
            Nome do arquivo pickle

        Returns
        -------
        object
            Objeto
        """
        file_content = self.get_file(folder, filename)
        return pickle.loads(file_content)

    def read_pickle_to_dataframe(
        self,
        folder: str,
        filename: str,
    ) -> pd.DataFrame:
        """
        Leitura de um pickle do sharepoint como dataframe.

        Parameters
        ----------
        folder : str
            Pasta de onde será lido o excel.
        filename : str
            Nome do arquivo excel a ser lido.

        Returns
        -------
        DataFrame

        """
        file_content = self.get_file(folder, filename)

        return pd.read_pickle(io.BytesIO(file_content))

    def read_df_from_excel(
        self,
        folder: str,
        filename: str,
        **kwargs,
    ) -> pd.DataFrame:
        """
        Leitura de um excel do sharepoint como dataframe.

        Parameters
        ----------
        folder : str
            Pasta de onde será lido o excel.
        filename : str
            Nome do arquivo excel a ser lido.
        **kwargs
            Parametros repassados para o método dataframe.read_excel

        Returns
        -------
        DataFrame

        """
        file_content = self.get_file(folder, filename)

        return pd.read_excel(io.BytesIO(file_content), engine="openpyxl", **kwargs)

    def read_sheetnames_from_excel(
        self,
        folder: str,
        filename: str,
    ) -> list[str]:
        """
        Obtêm os nomes das abas da planilha do excel

        Parameters
        ----------
        folder : str
            Pasta de onde será lido o excel.
        filename : str
            Nome do arquivo excel a ser lido.

        Returns
        -------
        list[str]
            Lista contendo os nomes das abas contidas no excel
        """
        return openpyxl.load_workbook(
            io.BytesIO(self.get_file(folder, filename))
        ).sheetnames

    def write_df_to_csv(
        self,
        dataframe: pd.DataFrame,
        folder: str,
        filename: str,
        **kwargs,
    ) -> None:
        """
        Salva o dataframe como csv no sharepoint.

        Parameters
        ----------
        dataframe : DataFrame
            Dataframe a ser salvo.
        folder : str
            Pasta onde será salvo.
        filename : str
            Nome do arquivo excel de destino.
        **kwargs
            Parametros repassados para o método dataframe.to_csv

        Returns
        -------
        None

        """
        # Cria um objeto binário
        bytes_io = io.BytesIO()
        # Salva o conteúdo neste objeto
        dataframe.to_csv(bytes_io, **kwargs)
        # Realiza o upload do conteúdo
        return self.upload_file(folder, bytes_io, filename)

    def write_df_to_excel(
        self,
        dataframe: pd.DataFrame,
        folder: str,
        filename: str,
        index: bool = False,
    ) -> None:
        """
        Salva o dataframe como excel no sharepoint.

        Parameters
        ----------
        dataframe : DataFrame
            Dataframe a ser salvo.
        folder : str
            Pasta onde será salvo.
        filename : str
            Nome do arquivo excel de destino.
        index : bool, optional
            Se o index da tabela será salvo.
            The default is False.

        Returns
        -------
        None

        """
        # Cria um objeto binário
        bytes_io = io.BytesIO()
        # Salva o conteúdo neste objeto
        dataframe.to_excel(bytes_io, engine="openpyxl", index=index)
        # Realiza o upload do conteúdo
        return self.upload_file(folder, bytes_io, filename)

    def delete_file(self,
                    folder: str,
                    filename: str):
        
        _folder = self.get_folder(folder)

        try:
            _folder.delete_file(filename)
        except ShareplumRequestError as err:
            _analyze_shareplum_err(err)